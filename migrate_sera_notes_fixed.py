"""Run Sera note migration from an exported Sera Jobs Report.

Safe development version for:
- rerunning ONLY jobs whose latest progress result is FAILED
- preserving one ServiceTitan Job Summary and appending additional Sera comments to it
- retrying transient ServiceTitan Add/Edit control timeouts

Usage:
  python migrate_sera_notes_fixed.py
  python migrate_sera_notes_fixed.py --last 4000
  python migrate_sera_notes_fixed.py --failed-only
"""
import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook

import migrate_sera_notes as migration

migration.MAX_JOBS = None


def parse_args():
    parser = argparse.ArgumentParser()
    group = parser.add_mutually_exclusive_group()
    group.add_argument(
        "--last",
        type=int,
        default=None,
        metavar="N",
        help="rerun only the last N jobs in Jobs Report order, ignoring progress status",
    )
    group.add_argument(
        "--failed-only",
        action="store_true",
        help="rerun only jobs whose latest progress-log result is FAILED",
    )
    args = parser.parse_args()
    if args.last is not None and args.last <= 0:
        parser.error("--last must be greater than 0")
    return args


def find_jobs_report():
    roots = [Path.cwd(), Path(__file__).resolve().parent, Path.home() / "Downloads", migration.APP_DATA]
    candidates, seen = [], set()
    for root in roots:
        if not root.exists():
            continue
        for path in root.glob("JobsReport*.xlsx"):
            try:
                resolved = path.resolve()
            except Exception:
                resolved = path
            if resolved in seen:
                continue
            seen.add(resolved)
            candidates.append(path)
    if not candidates:
        raise RuntimeError("No Sera JobsReport*.xlsx found. Put the Sera Jobs Report in Downloads or beside this script.")
    return max(candidates, key=lambda p: p.stat().st_mtime)


def load_jobs_report_ids(path):
    """Read unique numeric Job IDs while preserving spreadsheet row order."""
    print(f"Using Sera Jobs Report: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [str(value or "").strip() for value in next(rows)]
    except StopIteration:
        workbook.close()
        raise RuntimeError("Jobs Report is empty.")
    if "Job" not in headers:
        workbook.close()
        raise RuntimeError(f"Jobs Report does not contain the expected 'Job' column. Columns: {headers}")

    job_index = headers.index("Job")
    found, seen = [], set()
    invalid = 0
    for row in rows:
        if job_index >= len(row):
            continue
        job_id = migration.clean_id(row[job_index])
        if not job_id:
            continue
        if job_id.isdigit():
            if job_id not in seen:
                seen.add(job_id)
                found.append(job_id)
        else:
            invalid += 1
    workbook.close()

    if not found:
        raise RuntimeError("Jobs Report contained zero usable numeric job IDs.")
    print(f"Sera Jobs Report contains {len(found)} unique job IDs.")
    if invalid:
        print(f"WARNING: ignored {invalid} non-numeric Job value(s).")
    return found


def load_latest_progress_statuses():
    """Return the LAST recorded status for each job, not any historical status."""
    latest = {}
    path = migration.PROGRESS_LOG
    if not path.exists():
        return latest
    with path.open("r", newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            job_id = migration.clean_id(row.get("job_id"))
            status = str(row.get("status") or "").strip().upper()
            if job_id:
                latest[job_id] = status
    return latest


def install_job_discovery(last_n=None, failed_only=False):
    report_path = find_jobs_report()
    report_ids = load_jobs_report_ids(report_path)
    if len(report_ids) < 1000:
        raise RuntimeError(f"Jobs Report yielded only {len(report_ids)} jobs. Expected the full export; refusing bulk WRITE run.")

    if failed_only:
        latest = load_latest_progress_statuses()
        selected = [job_id for job_id in report_ids if latest.get(job_id) == "FAILED"]
        print(f"FAILED-ONLY MODE: {len(selected)} job(s) have latest status FAILED.")
        print("Those jobs will be retried regardless of earlier COMPLETE/NO_COMMENTS history.")
        migration.load_all_job_ids = lambda: list(selected)
        migration.load_completed_jobs = lambda: set()
        return

    if last_n is not None:
        selected = report_ids[-last_n:] if last_n < len(report_ids) else list(report_ids)
        print(f"RERUN RANGE MODE: selecting last {len(selected)} jobs from Jobs Report order.")
        print("Progress COMPLETE/NO_COMMENTS status will be ignored for this selected range.")
        migration.load_all_job_ids = lambda: list(selected)
        migration.load_completed_jobs = lambda: set()
        return

    local_loader = migration.load_all_job_ids

    def load_complete_job_ids():
        local = set(local_loader())
        report_set = set(report_ids)
        combined = report_set | local
        extra_local = local - report_set
        print(f"Job discovery: {len(report_ids)} from Jobs Report + {len(extra_local)} extra local-only = {len(combined)} unique jobs.")
        return sorted(combined, key=int, reverse=True)

    migration.load_all_job_ids = load_complete_job_ids


def comment_body(note_text):
    parts = note_text.split("\n\n", 1)
    return parts[1] if len(parts) == 2 else note_text


def page_contains_exact_manual_duplicate(page, note_text):
    body = comment_body(note_text)
    if not body:
        return False
    try:
        visible = page.locator("body").inner_text()
    except Exception:
        return False
    return body in visible


def already_present(page, note_text, marker, destination):
    if migration.page_contains_marker(page, marker):
        print(f"Duplicate detected on {destination}: migration marker already exists; skipping.")
        return True
    if page_contains_exact_manual_duplicate(page, note_text):
        print(f"Duplicate detected on {destination}: EXACT Sera message already exists; skipping.")
        return True
    return False


def wait_for_one(locator, description, timeout=30000):
    locator.first.wait_for(state="visible", timeout=timeout)
    count = locator.count()
    if count != 1:
        raise RuntimeError(f"Expected exactly one visible {description}, found {count}")
    return locator.first


def reload_and_wait(page):
    page.reload(wait_until="domcontentloaded", timeout=60000)
    page.wait_for_timeout(1800)


def find_summary_edit_button(page):
    """Locate the existing Job Summary edit control without relying on one generated CSS class."""
    # Best case: ServiceTitan exposes a summary-specific tracking id with edit semantics.
    tracked = page.locator('button[data-tracking-id*="summary"]:visible')
    tracked_matches = []
    for i in range(tracked.count()):
        button = tracked.nth(i)
        try:
            tracking = (button.get_attribute("data-tracking-id") or "").lower()
            text = " ".join(button.inner_text().split()).lower()
            html = (button.evaluate("el => el.outerHTML") or "").lower()
            if "jpm-job-add-button" in tracking or "summary-tab-content-button" in tracking:
                continue
            if "edit" in tracking or "edit" in text or "a-icon--edit" in html:
                tracked_matches.append(button)
        except Exception:
            pass
    if len(tracked_matches) == 1:
        return tracked_matches[0]

    # Fallback: an Edit icon/button whose nearby section text identifies it as Summary.
    edits = page.locator('button:visible:has(i.a-Icon--edit), button:visible:has-text("Edit")')
    nearby = []
    for i in range(edits.count()):
        button = edits.nth(i)
        try:
            context_text = button.evaluate(
                "el => { const p = el.closest('section, article, [class*=summary], [class*=Summary], [data-tracking-id]') || el.parentElement; return (p?.innerText || '').slice(0, 1200); }"
            )
            if "summary" in str(context_text or "").lower():
                nearby.append(button)
        except Exception:
            pass
    if len(nearby) == 1:
        return nearby[0]

    raise RuntimeError(
        "Existing Job Summary found, but its Edit control could not be identified safely. "
        "Need the existing-summary Edit button OuterHTML if ServiceTitan changed its markup."
    )


def open_job_summary_editor(page):
    """Open the single Job Summary for create OR edit, retrying transient UI delays."""
    for attempt in range(1, 4):
        try:
            add_button = page.locator('button[data-tracking-id="jpm-job-add-button"]:visible')
            if add_button.count() == 1:
                add_button.first.click()
                editor = page.locator('div[contenteditable="true"][class*="wysiwyg-editor"]:visible')
                return wait_for_one(editor, "job Summary editor")

            # No + Add Summary means a summary already exists. Edit that one instead.
            edit_button = find_summary_edit_button(page)
            edit_button.scroll_into_view_if_needed()
            edit_button.click()
            editor = page.locator('div[contenteditable="true"][class*="wysiwyg-editor"]:visible')
            return wait_for_one(editor, "job Summary editor")
        except Exception as exc:
            if attempt == 3:
                raise
            print(f"Job Summary editor attempt {attempt}/3 failed: {exc}; reloading and retrying...")
            reload_and_wait(page)
    raise RuntimeError("Could not open Job Summary editor")


def append_to_editor(editor, note_text):
    existing = editor.inner_text().rstrip()
    combined = note_text if not existing else f"{existing}\n\n{note_text}"
    editor.scroll_into_view_if_needed()
    editor.click()
    editor.focus()
    editor.evaluate(
        "(el, value) => { el.innerText = value; el.dispatchEvent(new InputEvent('input', {bubbles:true, inputType:'insertText', data:value})); }",
        combined,
    )


def add_job_summary_exact(page, note_text, marker):
    """Maintain ONE ServiceTitan summary and append each missing Sera comment below it."""
    if already_present(page, note_text, marker, "job"):
        return "SKIPPED"
    if migration.DRY_RUN:
        return "DRY_RUN"

    editor = open_job_summary_editor(page)

    # Re-check after opening editor because the full existing summary may only become
    # visible in edit mode. This prevents appending a manual duplicate hidden in display mode.
    try:
        existing_editor_text = editor.inner_text()
    except Exception:
        existing_editor_text = ""
    body = comment_body(note_text)
    if marker in existing_editor_text or (body and body in existing_editor_text):
        print("Duplicate detected inside existing Job Summary editor; skipping append.")
        # Escape edit mode without changing content when possible.
        page.keyboard.press("Escape")
        return "SKIPPED"

    append_to_editor(editor, note_text)
    print("Appended Sera comment to the single ServiceTitan Job Summary.")

    submit = page.locator('button[data-tracking-id="jpm-job-summary-tab-content-button"]:visible')
    submit = wait_for_one(submit, "job Summary save button")
    submit.click()
    page.wait_for_timeout(1500)
    if not migration.page_contains_marker(page, marker):
        raise RuntimeError("Job Summary save was clicked, but migrated note marker is not visible")
    return "SUCCESS"


def add_customer_note_safe(page, note_text, marker):
    if already_present(page, note_text, marker, "customer"):
        return "SKIPPED"
    if migration.DRY_RUN:
        return "DRY_RUN"

    last_error = None
    for attempt in range(1, 4):
        try:
            add_button = page.locator('button[data-tracking-id="crm-notes-add-note-button"]:visible')
            add_button = wait_for_one(add_button, "customer Add Note button", timeout=30000)
            add_button.click()

            note_box = page.locator('textarea[placeholder="Leave a note..."][data-anvil-component="TextArea"]:visible')
            note_box = wait_for_one(note_box, "customer note textarea")
            note_box.scroll_into_view_if_needed()
            note_box.click()
            note_box.focus()
            note_box.fill(note_text)

            submit = page.locator('button[data-tracking-id="add-note-button"]:visible')
            submit = wait_for_one(submit, "Add Note submit button")
            submit.click()
            page.wait_for_timeout(1500)
            if not migration.page_contains_marker(page, marker):
                raise RuntimeError("Customer Add Note was clicked, but migrated note marker is not visible")
            return "SUCCESS"
        except Exception as exc:
            last_error = exc
            if attempt == 3:
                break
            print(f"Customer Add Note attempt {attempt}/3 failed: {exc}; reloading and retrying...")
            reload_and_wait(page)
            if already_present(page, note_text, marker, "customer"):
                return "SKIPPED"

    raise RuntimeError(f"Customer Add Note failed after 3 attempts: {last_error}")


migration.add_job_summary = add_job_summary_exact
migration.add_customer_note = add_customer_note_safe

if __name__ == "__main__":
    args = parse_args()
    install_job_discovery(args.last, args.failed_only)
    migration.main()
