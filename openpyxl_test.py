from openpyxl import load_workbook
from pathlib import Path

path = Path("exports") / "CustomerContactReport-2026-07-22-58a68e.xlsx"

print("Opening:", path)

wb = load_workbook(path, read_only=True)

print("Success!")
print(wb.sheetnames)

wb.close()