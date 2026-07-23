from openpyxl import load_workbook


class CustomerLookup:

    def __init__(self, excel_file):

        self.customers = {}

        workbook = load_workbook(
            excel_file,
            read_only=True,
            data_only=True
        )

        sheet = workbook.active

        headers = [
            str(cell.value).strip()
            if cell.value else ""
            for cell in sheet[1]
        ]

        header = {
            name: index
            for index, name in enumerate(headers)
        }

        for row in sheet.iter_rows(
            min_row=2,
            values_only=True
        ):

            row = list(row)

            while len(row) < len(headers):
                row.append(None)

            legacy = row[header["Legacy ID"]]

            if legacy is None:
                continue

            legacy = str(legacy).strip()

            self.customers[legacy] = {
                "customer_id": str(row[header["ID"]]).strip(),
                "customer_name": str(row[header["Name"]]).strip()
            }

        workbook.close()

    def find(self, legacy_id):

        return self.customers.get(
            str(legacy_id)
        )