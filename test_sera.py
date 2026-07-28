from live_migration import LiveMigration
from playwright.sync_api import sync_playwright
from pathlib import Path
from openpyxl import load_workbook
from resume_tracker import ResumeTracker
from datetime import datetime, timedelta
from progress import progress
from cancel import cancel
import csv
import time
import re


progress.log("Starting test_sera...")


# ============================================================
# SETTINGS
# ============================================================

EXCEL_FILE = Path("exports") / "CustomerContactReport-2026-07-22-58a68e.xlsx"

BASE_URL = "https://grmetro.sera.tech/customers"

DOWNLOAD_FOLDER = Path("sera_media")

DOWNLOAD_FOLDER.mkdir(exist_ok=True)

UPLOADED_FOLDER = Path("uploaded_media")

UPLOADED_FOLDER.mkdir(exist_ok=True)

LOG_FILE = Path("download_log.csv")

# TEST FIRST
#TEST_MODE = True

#TEST_CUSTOMER_COUNT = 1

RUNTIME_EXCEL_FILE = None

# ============================================================
# HELPER FUNCTIONS
# ============================================================

def clean_filename(filename):

    if not filename:
        return None

    filename = str(filename).strip()

    filename = re.sub(
        r'[<>:"/\\|?*]',
        '-',
        filename
    )

    filename = filename.rstrip(" .")

    return filename


def get_extension(content_type):

    if not content_type:
        return ""

    content_type = content_type.split(";")[0].strip()

    extensions = {

        "image/jpeg": ".jpg",

        "image/jpg": ".jpg",

        "image/png": ".png",

        "image/gif": ".gif",

        "image/webp": ".webp",

        "application/pdf": ".pdf",

        "application/msword": ".doc",

        "application/vnd.openxmlformats-officedocument.wordprocessingml.document": ".docx",

        "application/vnd.ms-excel": ".xls",

        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",

    }

    return extensions.get(content_type, "")


def get_job_number(image):

    """
    Finds the Job # associated with this media item.
    """

    # Start with the image's nearest reasonable container.
    # We walk upward through parent elements and look for text
    # containing "Job #".
    for level in range(1, 7):

        try:

            container = image.locator(
                ".." * level
            )

            text = container.inner_text(
                timeout=1000
            )

            match = re.search(
                r'Job\s*#\s*(\d+)',
                text,
                re.IGNORECASE
            )

            if match:

                return match.group(1)

        except:

            pass


    # If no job number is found, place it in an
    # Unassigned folder instead of guessing.

    return "Unassigned"


def initialize_log():

    if not LOG_FILE.exists():

        with open(
            LOG_FILE,
            "w",
            newline="",
            encoding="utf-8"
        ) as file:

            writer = csv.writer(file)

            writer.writerow([

                "customer_id",

                "job_number",

                "filename",

                "saved_path",

                "status",

                "error"

            ])


def write_log(
    customer_id,
    job_number,
    filename,
    saved_path,
    status,
    error=""
):

    with open(
        LOG_FILE,
        "a",
        newline="",
        encoding="utf-8"
    ) as file:

        writer = csv.writer(file)

        writer.writerow([

            customer_id,

            job_number,

            filename,

            saved_path,

            status,

            error

        ])


# ============================================================
# READ SERA CUSTOMER IDS
# ============================================================

def run(
    workbook=None,
    limit=0
):

    
    
    completed = False

    progress.log("Entered run()")

    global RUNTIME_EXCEL_FILE

    RUNTIME_EXCEL_FILE = workbook

    progress.log("Reading customer list...")

    workbook_path = workbook or EXCEL_FILE

    progress.log(f"Workbook: {workbook_path}")
    print("Absolute: " + str(Path(workbook_path).resolve()))
    print("Exists: " + str(Path(workbook_path).exists()))

    #from pathlib import Path

    print("Absolute: " + str(Path(RUNTIME_EXCEL_FILE or EXCEL_FILE).resolve()))
    print("Exists: " + str(Path(RUNTIME_EXCEL_FILE or EXCEL_FILE).exists()))

    workbook_path = workbook or EXCEL_FILE

    workbook = load_workbook(
        workbook_path,
        read_only=True
    )

    start_time = time.time()

    sheet = workbook.active

    customer_ids = []

    tracker = ResumeTracker()

    uploaded_count = 0
    skipped_count = 0
    failed_count = 0

    last_customer = tracker.load()

    if last_customer:

        progress.log(
            f"Resuming after customer "
            f"{last_customer}"
        )

    resume = last_customer is None

    skip_current = last_customer is not None

    # Column A = Id

    for row in sheet.iter_rows(

        min_row=2,

        max_col=1,

        values_only=True

    ):

        customer_id = row[0]

        if customer_id:

            customer_ids.append(
                str(customer_id).strip()
            )


    progress.log(
        f"Found {len(customer_ids)} customers."
    )


    if limit > 0:

        customer_ids = customer_ids[15:17]

        progress.log(
            f"TEST MODE: Processing "
            f"{len(customer_ids)} customers."
        )

    else:

        progress.log(
            f"FULL MODE: Processing "
            f"{len(customer_ids)} customers."
        )


    initialize_log()


    # ============================================================
    # OPEN SERA
    # ============================================================

    with sync_playwright() as p:

        

        try:

            progress.log("Entered Playwright")

            context = p.chromium.launch_persistent_context(

                "sera_browser_profile",

                headless=False

            )


            page = (

                context.pages[0]

                if context.pages

                else context.new_page()

            )


            # ========================================================
            # PROCESS CUSTOMERS
            # ========================================================

            #migration = LiveMigration(p)

            progress.log("Created LiveMigration")

            progress.log("Connected to Sera")

            progress.log("Starting customer loop")

            for customer_number, customer_id in enumerate(

                customer_ids,

                start=1

            ):

                if cancel.is_cancelled():

                    progress.log("Download cancelled.")

                    break

                progress.log(f"Processing customer {customer_id}")

                if cancel.is_cancelled():

                    progress.log("Migration cancelled.")

                    break
                progress.progress(
                    customer_number,
                    len(customer_ids)
                )

                if customer_number > 0:

                    elapsed = time.time() - start_time

                    #progress.elapsed(
                    #    str(timedelta(seconds=int(elapsed)))
                    #)

                    average = elapsed / customer_number

                    remaining = len(customer_ids) - customer_number

                    eta = datetime.now() + timedelta(
                        seconds=average * remaining
                    )

                    progress.eta(
                        eta.strftime("%I:%M %p")
        )
                #
                # Resume support
                #
                if not resume:

                    if str(customer_id) == last_customer:
                        resume = True

                    continue

                #
                # Skip the customer we already completed
                #

                if skip_current:

                    skip_current = False

                    continue

                progress.log("\n" + "=" * 60)

                progress.log(

                    f"[{customer_number}/"
                    f"{len(customer_ids)}] "
                    f"Customer {customer_id}"

                )

                progress.log("=" * 60)

                progress.customer(customer_id)

                progress.log(f"Customer: {customer_id}")

                customer_folder = (

                    DOWNLOAD_FOLDER /

                    f"Customer_{customer_id}"

                )

                customer_folder.mkdir(

                    exist_ok=True

                )


                url = (

                    f"{BASE_URL}/"

                    f"{customer_id}"

                    f"?tab=c_Media+and+Documents"

                )


                try:

                    progress.action("Opening customer...")

                    page.goto(

                        url,

                        wait_until="domcontentloaded",

                        timeout=60000

                    )


                    page.wait_for_timeout(1500)


                    images = page.locator("img")


                    image_count = images.count()


                    progress.log(

                        f"Found {image_count} "
                        f"image elements"

                    )

                    real_media_number = 0

                    customer_completed = True

                    for image_number in range(image_count):

                        filename = ""

                        if cancel.is_cancelled():

                            progress.log("Migration cancelled.")

                            break


                        try:

                            image = images.nth(

                                image_number

                            )


                            filename = (

                                image.get_attribute(

                                    "alt"

                                )

                            )


                            # Skip fake UI image

                            if (

                                filename

                                and

                                filename.strip().lower()

                                == "membership icon"

                            ):

                                progress.log(

                                    "  Skipping "
                                    "Membership Icon"

                                )

                                continue


                            real_media_number += 1


                            # ------------------------------------------------
                            # FIND JOB NUMBER
                            # ------------------------------------------------

                            progress.action("Finding job...")

                            job_number = get_job_number(

                                image

                            )


                            progress.log(

                                f"  Job: #{job_number}"

                            )

                            progress.job(job_number)


                            # ------------------------------------------------
                            # CREATE JOB FOLDER
                            # ------------------------------------------------

                            job_folder = (

                                customer_folder /

                                f"Job_{job_number}"

                            )


                            job_folder.mkdir(

                                exist_ok=True

                            )


                            # ------------------------------------------------
                            # OPEN MEDIA
                            # ------------------------------------------------

                            image.click()


                            page.wait_for_timeout(

                                700

                            )


                            download_link = (

                                page.get_by_role(

                                    "link",

                                    name="Download"

                                )

                            )


                            if (

                                download_link.count()

                                == 0

                            ):

                                progress.log(

                                    "  Could not find "
                                    "Download link"

                                )


                                write_log(

                                    customer_id,

                                    job_number,

                                    filename,

                                    "",

                                    "Skipped",

                                    "Download link not found"

                                )


                                page.keyboard.press(

                                    "Escape"

                                )


                                continue


                            href = (

                                download_link

                                .first

                                .get_attribute(

                                    "href"

                                )

                            )


                            if not href:

                                progress.log(

                                    "  Download link had "
                                    "no URL"

                                )

                                continue


                            # ------------------------------------------------
                            # DOWNLOAD FILE
                            # ------------------------------------------------

                            progress.action("Downloading file...")
                            
                            response = (

                                page.request.get(

                                    href,

                                    timeout=60000

                                )

                            )


                            if response.status != 200:

                                progress.log(

                                    f"  Download failed: "
                                    f"HTTP {response.status}"

                                )

                                write_log(

                                    customer_id,

                                    job_number,

                                    filename,

                                    "",

                                    "Failed",

                                    f"HTTP {response.status}"

                                )

                                continue


                            # ------------------------------------------------
                            # FILENAME
                            # ------------------------------------------------

                            filename = clean_filename(

                                filename

                            )


                            if not filename:

                                filename = (

                                    f"media_"

                                    f"{real_media_number}"

                                )


                            # Add extension if needed

                            if not Path(

                                filename

                            ).suffix:

                                extension = get_extension(

                                    response.headers.get(

                                        "content-type",

                                        ""

                                    )

                                )

                                filename += extension


                            # ------------------------------------------------
                            # SAVE FILE
                            # ------------------------------------------------

                            file_path = (

                                job_folder /

                                filename

                            )

                            uploaded_path = (
                                UPLOADED_FOLDER
                                / file_path.relative_to(DOWNLOAD_FOLDER)
                            )

                            #
                            # Already migrated?
                            #

                            if uploaded_path.exists():

                                progress.log(f"✓ Already migrated: {filename}")

                                skipped_count += 1

                                progress.skipped(skipped_count)

                                page.keyboard.press("Escape")

                                page.wait_for_timeout(400)

                                continue


                            # Avoid overwriting

                            #
                            # Already downloaded?
                            #
                            if file_path.exists():

                                progress.log(f"Already downloaded: {filename}")

                                progress.file(filename)

                                progress.action("Already downloaded")

                                page.keyboard.press("Escape")

                                page.wait_for_timeout(400)

                                continue


                            file_path.write_bytes(

                                response.body()

                            )

                            progress.file(filename)

                            progress.log(

                                f"  SAVED: {file_path}"

                            )


                            write_log(

                                customer_id,

                                job_number,

                                filename,

                                str(file_path),

                                "Downloaded"

                            )


                            # Close viewer

                            page.keyboard.press(

                                "Escape"

                            )


                            page.wait_for_timeout(

                                400

                            )


                        except Exception as error:

                            customer_completed = False

                            progress.log(

                                f"  ERROR: {error}"

                            )


                            write_log(

                                customer_id,

                                "Unknown",

                                filename,

                                "",

                                "Error",

                                str(error)

                            )


                            try:

                                page.keyboard.press(

                                    "Escape"

                                )

                            except:

                                pass


                    page.wait_for_timeout(1000)

                    if customer_completed:

                        tracker.save(customer_id)

                except Exception as e:

                    import traceback

                    traceback.print_exc()

                    progress.log(f"Customer failed: {customer_id}")
                    progress.log(str(e))

                    failed_count += 1

                    continue

            completed = True

            progress.log("\n" + "=" * 60)

            progress.log("TEST COMPLETE")

            progress.action("Finished")

            progress.log("=" * 60)

            progress.log(

                f"Files saved in: "

                f"{DOWNLOAD_FOLDER.absolute()}"

            )

            progress.log(

                f"Log saved in: "

                f"{LOG_FILE.absolute()}"

            )

            if completed:
                tracker.clear()

            #migration.close()

            context.close()

        except Exception:
            import traceback
            traceback.print_exc()
            progress.log("Fatal error while downloading.")
            progress.log(str(traceback.format_exc()))

if __name__ == "__main__":
    run()