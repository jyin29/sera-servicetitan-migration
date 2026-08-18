import csv
import os
import sqlite3
from pathlib import Path

from playwright.sync_api import sync_playwright

from sera.notes import get_or_open_sera_page, load_job_data
from servicetitan.browser import connect, wait_for_servicetitan
from servicetitan.job_search import JobSearcher
from servicetitan.customer_search import CustomerSearcher

APP_DATA = Path(os.getenv("LOCALAPPDATA")) / "Sera ServiceTitan Migration"
MIGRATION_LOG = APP_DATA / "migration_log.csv"
MEDIA_ROOTS = [APP_DATA / "sera_media", Path(__file__).resolve().parent / "sera_media"]
MIGRATION_DATABASES = [APP_DATA / "migration.db", Path(__file__).resolve().parent / "database" / "migration.db"]

# Customer matching exports used earlier in this migration project.
MATCHING_FILES = [
    APP_DATA / "Sera_ServiceTitan_Matching.xlsx",
    Path(__file__).resolve().parent / "Sera_ServiceTitan_Matching.xlsx",
]

JOB_IDS = ["6505724"]
DRY_RUN = True


def clean_id(value):
    value = str(value or "").strip()
    if value.endswith(".0") and value[:-2].isdigit():
        value = value[:-2]
    return value


def add_mapping(mapping, sources, job_id, legacy_id, source):
    job_id, legacy_id = clean_id(job_id), clean_id(legacy_id)
    if not job_id or not legacy_id or legacy_id.lower() == "none":
        return
    if job_id not in mapping:
        mapping[job_id] = legacy_id
        sources[job_id] = source


def load_job_to_customer_map():
    mapping, sources = {}, {}

    for db_path in MIGRATION_DATABASES:
        if not db_path.exists():
            continue
        try:
            connection = sqlite3.connect(str(db_path))
            cursor = connection.cursor()
            cursor.execute("SELECT jobs.sera_job_number, customers.legacy_id FROM jobs JOIN customers ON customers.id = jobs.customer_id WHERE jobs.sera_job_number IS NOT NULL AND customers.legacy_id IS NOT NULL")
            for job_id, legacy_id in cursor.fetchall():
                add_mapping(mapping, sources, job_id, legacy_id, f"database {db_path}")
            connection.close()
        except Exception as exc:
            print(f"Could not read {db_path}: {exc}")

    if MIGRATION_LOG.exists():
        with MIGRATION_LOG.open("r", newline="", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                add_mapping(mapping, sources, row.get("Job Number"), row.get("Legacy ID"), f"migration log {MIGRATION_LOG}")

    for media_root in MEDIA_ROOTS:
        if not media_root.exists():
            continue
        for customer_dir in media_root.glob("Customer_*"):
            legacy_id = customer_dir.name.replace("Customer_", "", 1).strip()
            for job_dir in customer_dir.glob("Job_*"):
                add_mapping(mapping, sources, job_dir.name.replace("Job_", "", 1), legacy_id, f"media folder {job_dir}")

    return mapping, sources


def find_matching_workbook():
    for path in MATCHING_FILES:
        if path.exists():
            return path
    # Also search common project/app-data locations without requiring a fixed filename location.
    roots = [APP_DATA, Path(__file__).resolve().parent, Path.home() / "Downloads", Path.home() / "Documents"]
    for root in roots:
        if not root.exists():
            continue
        matches = list(root.glob("Sera_ServiceTitan_Matching*.xlsx"))
        if matches:
            return matches[0]
    return None


def load_sera_customer_crosswalk():
    workbook = find_matching_workbook()
    if workbook is None:
        return {}, {}, None

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("WARNING: openpyxl is not installed; cannot read Sera_ServiceTitan_Matching.xlsx")
        return {}, {}, workbook

    by_sera_id, names = {}, {}
    wb = load_workbook(workbook, read_only=True, data_only=True)

    preferred = [name for name in ["Matched", "All Matches"] if name in wb.sheetnames]
    sheet_names = preferred or wb.sheetnames

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            header = [str(v or "").strip() for v in next(rows)]
        except StopIteration:
            continue

        columns = {name: i for i, name in enumerate(header)}
        if "Sera ID" not in columns or "ServiceTitan Legacy ID" not in columns:
            continue

        for row in rows:
            sera_id = clean_id(row[columns["Sera ID"]] if columns["Sera ID"] < len(row) else "")
            legacy_id = clean_id(row[columns["ServiceTitan Legacy ID"]] if columns["ServiceTitan Legacy ID"] < len(row) else "")
            if not sera_id or not legacy_id:
                continue
            by_sera_id.setdefault(sera_id, legacy_id)
            if "Sera Customer Name" in columns and columns["Sera Customer Name"] < len(row):
                names.setdefault(sera_id, str(row[columns["Sera Customer Name"]] or "").strip())

    wb.close()
    return by_sera_id, names, workbook


def format_note(job_id, comment):
    stamp = comment.get("timestamp") or " ".join(part for part in [comment.get("date", ""), comment.get("time", "")] if part)
    author = comment.get("author") or "Unknown"
    marker = f"[Migrated from Sera | Job {job_id} | {stamp} | {author}]"
    return marker, f"{marker}\n\n{comment['text'].strip()}"


def page_contains_marker(page, marker):
    try:
        return page.get_by_text(marker, exact=False).count() > 0
    except Exception:
        return marker in page.locator("body").inner_text()


def find_single_visible_editor(page):
    candidates = page.locator("textarea:visible, [contenteditable='true']:visible, input[type='text']:visible")
    usable = []
    for i in range(candidates.count()):
        item = candidates.nth(i)
        try:
            if item.is_enabled() and item.is_visible():
                usable.append(item)
        except Exception:
            pass
    if len(usable) != 1:
        raise RuntimeError(f"Expected exactly one visible note editor after clicking Add, found {len(usable)}")
    return usable[0]


def fill_editor(editor, text):
    tag = editor.evaluate("el => el.tagName.toLowerCase()")
    if tag in {"textarea", "input"}:
        editor.fill(text)
    else:
        editor.click()
        editor.evaluate("(el, value) => { el.innerText = value; el.dispatchEvent(new InputEvent('input', {bubbles:true, inputType:'insertText', data:value})); }", text)


def click_single_save_button(page):
    buttons = page.locator("button:visible")
    matches = []
    for i in range(buttons.count()):
        button = buttons.nth(i)
        try:
            if button.is_enabled() and " ".join(button.inner_text().split()).strip().lower() in {"save", "add", "add note", "add summary", "submit"}:
                matches.append(button)
        except Exception:
            pass
    if len(matches) != 1:
        raise RuntimeError(f"Expected exactly one visible Save/Add button, found {len(matches)}")
    matches[0].click()


def add_job_summary(page, note_text, marker):
    if page_contains_marker(page, marker):
        print("Already present on job; skipping duplicate.")
        return "SKIPPED"
    add_button = page.locator('button[data-tracking-id="jpm-job-add-button"]')
    add_button.wait_for(state="visible", timeout=10000)
    if DRY_RUN:
        print("DRY RUN: would click + Add Summary and add:")
        print(note_text)
        return "DRY_RUN"
    add_button.click()
    page.wait_for_timeout(500)
    fill_editor(find_single_visible_editor(page), note_text)
    click_single_save_button(page)
    page.wait_for_timeout(1200)
    if not page_contains_marker(page, marker):
        raise RuntimeError("Job summary save was clicked, but migrated note marker is not visible")
    return "SUCCESS"


def add_customer_note(page, note_text, marker):
    if page_contains_marker(page, marker):
        print("Already present on customer; skipping duplicate.")
        return "SKIPPED"
    add_button = page.locator('button[data-tracking-id="crm-notes-add-note-button"]')
    add_button.wait_for(state="visible", timeout=10000)
    if DRY_RUN:
        print("DRY RUN: would click Add Note and add:")
        print(note_text)
        return "DRY_RUN"
    add_button.click()
    page.wait_for_timeout(500)
    fill_editor(find_single_visible_editor(page), note_text)
    click_single_save_button(page)
    page.wait_for_timeout(1200)
    if not page_contains_marker(page, marker):
        raise RuntimeError("Customer note save was clicked, but migrated note marker is not visible")
    return "SUCCESS"


def main():
    old_job_map, old_sources = load_job_to_customer_map()
    sera_crosswalk, sera_names, workbook = load_sera_customer_crosswalk()

    print(f"Customer crosswalk: {workbook or 'NOT FOUND'}")
    print(f"Sera customers loaded from crosswalk: {len(sera_crosswalk)}")

    with sync_playwright() as p:
        browser, context = connect(p)
        st_page = wait_for_servicetitan(context)
        sera_page = get_or_open_sera_page(context)
        job_search = JobSearcher(st_page)
        customer_search = CustomerSearcher(st_page)

        total_notes = completed = skipped = 0
        failures = []

        for job_id in JOB_IDS:
            print("=" * 80)
            print(f"SERA NOTE MIGRATION: job {job_id}")
            print(f"Mode: {'DRY RUN' if DRY_RUN else 'WRITE'}")
            print("=" * 80)

            try:
                job_data = load_job_data(sera_page, job_id)
            except Exception as exc:
                failures.append((job_id, f"Could not load Sera job: {exc}"))
                print(f"FAILED loading Sera job: {exc}")
                continue

            sera_customer_id = clean_id(job_data["sera_customer_id"])
            sera_customer_name = job_data["customer_name"]
            comments = job_data["comments"]
            total_notes += len(comments)

            # PRIMARY mapping: actual customer link on the actual Sera job page.
            legacy_id = sera_crosswalk.get(sera_customer_id)
            mapping_source = f"Sera job customer link + {workbook}" if legacy_id else None

            # Backup only: old media/database-derived job mapping.
            if not legacy_id:
                legacy_id = old_job_map.get(job_id)
                mapping_source = old_sources.get(job_id) if legacy_id else None

            print(f"Sera customer: {sera_customer_name}")
            print(f"Sera customer ID: {sera_customer_id}")
            print(f"ServiceTitan Legacy ID: {legacy_id or 'not found'}")
            if mapping_source:
                print(f"Mapping source: {mapping_source}")
            if sera_customer_id in sera_names and sera_names[sera_customer_id]:
                print(f"Crosswalk customer name: {sera_names[sera_customer_id]}")
            print(f"Found {len(comments)} Sera comment(s)")

            if not comments:
                continue

            st_page.bring_to_front()
            try:
                job_found = job_search.open_job(job_id)
            except Exception as exc:
                print(f"Job search error: {exc}")
                job_found = False

            destination = "JOB" if job_found else None
            if not job_found:
                if not legacy_id:
                    failures.append((job_id, f"No Legacy ID for Sera customer {sera_customer_id} ({sera_customer_name})"))
                    print("Cannot fall back to customer because the Sera customer is not mapped to a ServiceTitan Legacy ID.")
                    continue
                print(f"Job not found; falling back to customer Legacy ID {legacy_id}...")
                try:
                    if customer_search.open_customer(legacy_id):
                        destination = "CUSTOMER"
                    else:
                        failures.append((job_id, f"ServiceTitan customer {legacy_id} not found"))
                        continue
                except Exception as exc:
                    failures.append((job_id, f"Customer search error: {exc}"))
                    continue

            for number, comment in enumerate(comments, start=1):
                marker, note_text = format_note(job_id, comment)
                print("-" * 80)
                print(f"Comment {number}/{len(comments)} -> {destination}")
                print(marker)
                try:
                    result = add_job_summary(st_page, note_text, marker) if destination == "JOB" else add_customer_note(st_page, note_text, marker)
                    if result == "SKIPPED":
                        skipped += 1
                    else:
                        completed += 1
                except Exception as exc:
                    print(f"FAILED comment {number}: {exc}")
                    failures.append((job_id, f"Comment {number}: {exc}"))
                    break

        print("=" * 80)
        print("NOTE MIGRATION COMPLETE")
        print(f"Sera comments found: {total_notes}")
        print(f"Completed/would complete: {completed}")
        print(f"Already present/skipped: {skipped}")
        print(f"Failures: {len(failures)}")
        for job_id, reason in failures:
            print(f"- {job_id}: {reason}")
        print("=" * 80)
        input("Press Enter to exit...")


if __name__ == "__main__":
    main()
