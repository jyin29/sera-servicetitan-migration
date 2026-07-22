from openpyxl import load_workbook


class CustomerExcelReader:

    def __init__(self, excel_file):
        self.excel_file = excel_file

    def get_customer_ids(self):

        print("Opening:", self.excel_file)

        print(repr(self.excel_file))
        print(type(self.excel_file))

        workbook = load_workbook(
            filename=self.excel_file,
            read_only=True,
            #data_only=True,
        )

        sheet = workbook.active

        ids = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                ids.append(str(row[0]))

        workbook.close()

        return ids